import os
import openpyxl
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, jsonify

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('daily_expense.html')

@app.route('/daily-expense', methods=['GET', 'POST'])
def daily_expense():
    if request.method == 'POST':
        amount = request.form['amount']
        description = request.form['description']
        date = request.form['date']
        response = submit_expense(amount, description, date)
        return jsonify(response=response)
    return render_template('daily_expense.html')

@app.route('/document-scan', methods=['GET', 'POST'])
def document_scan():
    if request.method == 'POST':
        pan_number = request.form['pan_number']
        file = request.files['document_file']
        base64_file = file.read().decode('base64')
        response = upload_document(pan_number, base64_file)
        return jsonify(response=response)
    return render_template('document_scan.html')

@app.route('/call-log', methods=['GET', 'POST'])
def call_log():
    if request.method == 'POST':
        file = request.files['call_log_file']
        file_path = os.path.join('/tmp', file.filename)
        file.save(file_path)
        response = upload_call_log(file_path)
        return jsonify(response=response)
    return render_template('call_log.html')

@app.route('/reports', methods=['GET', 'POST'])
def reports():
    if request.method == 'POST':
        report_type = request.form['report_type']
        response = generate_report(report_type)
        return jsonify(response=response)
    return render_template('reports.html')

def submit_expense(amount, description, date):
    try:
        date_obj = datetime.strptime(date, "%Y-%m-%d")
        month = date_obj.strftime("%B")
        year = date_obj.year
        
        dir_path = os.path.join("D:", f"{month}_{year}")
        if not os.path.exists(dir_path):
            os.makedirs(dir_path)
        filename = os.path.join(dir_path, f"expenses_{date}.xlsx")
        
        if os.path.exists(filename):
            workbook = openpyxl.load_workbook(filename)
        else:
            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)

        if "Expenses" not in workbook.sheetnames:
            sheet = workbook.create_sheet(title='Expenses')
            sheet.append(['Amount', 'Description', 'Date'])
        else:
            sheet = workbook['Expenses']

        sheet.append([amount, description, date])
        workbook.save(filename)

        return "Expense saved successfully."
    except Exception as e:
        return f"An error occurred: {e}"

def generate_report(report_type):
    try:
        report_dir = os.path.join("D:", "Reports")
        if not os.path.exists(report_dir):
            os.makedirs(report_dir)

        filename = os.path.join(report_dir, f"{report_type}_report.txt")
        with open(filename, "w") as file:
            file.write(f"Report: {report_type.capitalize()} Report\n")
            file.write(f"Generated on: {datetime.now()}\n")
            file.write("This is a placeholder report.")

        return f"{report_type.capitalize()} report generated successfully at {filename}."
    except Exception as e:
        return f"An error occurred: {e}"
    
def upload_document(pan_number, base64_file):
    try:
        import base64
        file_data = base64.b64decode(base64_file)
        
        dir_path = os.path.join("D:", "PAN_Cards")
        if not os.path.exists(dir_path):
            os.makedirs(dir_path)

        destination_path = os.path.join(dir_path, f"{pan_number}.pdf")
        with open(destination_path, "wb") as file:
            file.write(file_data)

        return f"Document saved successfully as {pan_number}.pdf."
    except Exception as e:
        return f"An error occurred: {e}"

def upload_call_log(file_path):
    try:
        dir_path = os.path.join("D:", "Call_Logs")
        if not os.path.exists(dir_path):
            os.makedirs(dir_path)

        destination_path = os.path.join(dir_path, os.path.basename(file_path))
        os.rename(file_path, destination_path)

        return f"Call log saved successfully as {os.path.basename(file_path)}."
    except Exception as e:
        return f"An error occurred: {e}"

if __name__ == '__main__':
    app.run(debug=True)
